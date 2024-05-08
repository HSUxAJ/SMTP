import sys
import smtplib
from email.mime.text import MIMEText
from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QFileDialog, QLabel, QTextEdit, QVBoxLayout, QDialog
from PyQt5.QtCore import pyqtSlot
import openpyxl
import logging
import os

# 創建logger
logger = logging.getLogger('email_app')
logger.setLevel(logging.DEBUG)

# 將logger存進文檔
file_handler = logging.FileHandler('email_app.log')
file_handler.setLevel(logging.DEBUG)

# 在cmd顯示logger
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.DEBUG)

formatter = logging.Formatter('%(asctime)s-%(name)s-%(levelname)s-%(message)s')

file_handler.setFormatter(formatter)
console_handler.setFormatter(formatter)

logger.addHandler(file_handler)
logger.addHandler(console_handler)

class EditMessageDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle('編輯郵件內容')
        self.setGeometry(600, 100, 800, 800)

        layout = QVBoxLayout()

        self.text_edit = QTextEdit(self)
        layout.addWidget(self.text_edit)

        save_button = QPushButton('儲存', self)
        save_button.clicked.connect(self.accept)
        layout.addWidget(save_button)

        self.setLayout(layout)

    def set_text(self, text):
        self.text_edit.setHtml(text)

    def get_text(self):
        return self.text_edit.toHtml()

class EmailWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setGeometry(100, 100, 300, 400)
        self.setWindowTitle('發送email')

        self.import_status_label = QLabel('尚未匯入 Excel 檔案', self)
        self.import_status_label.setGeometry(50, 30, 300, 30)

        import_button = QPushButton('匯入Excel', self)
        import_button.setGeometry(50, 70, 200, 30)
        import_button.clicked.connect(self.import_file)

        send_button = QPushButton('傳送email', self)
        send_button.setGeometry(50, 110, 200, 30)
        send_button.clicked.connect(self.send_email_button_clicked)

        self.text_edit = QTextEdit(self)
        self.text_edit.setGeometry(0, 0, 0, 0)

        self.mail_format_lable = QLabel('尚未選擇信件編輯模板', self)
        self.mail_format_lable.setGeometry(50, 190, 300, 30)

        report_button = QPushButton('1. 回報駕駛', self)
        report_button.setGeometry(50, 230, 200, 30)
        report_button.clicked.connect(self.send_report_email_button_clicked)

        welcome_button_ch = QPushButton('2. 歡迎信(中文)', self)
        welcome_button_ch.setGeometry(50, 270, 200, 30)
        welcome_button_ch.clicked.connect(self.send_welcome_ch_email_button_clicked)

        welcome_button_eng = QPushButton('3. 歡迎信(英文)', self)
        welcome_button_eng.setGeometry(50, 310, 200, 30)
        welcome_button_eng.clicked.connect(self.send_welcome_eng_email_button_clicked)
        
        self.mail_format = []
        self.mail_type = 0
        all_files = os.listdir("mail_format/")
        for file_name in all_files:
            file_path = f'mail_format\{file_name}'
            with open(file_path, 'r', encoding="utf-8") as f:
                lines = f.readlines()
            content = ''
            for line in lines:
                content += line
            self.mail_format.append(content)

        # 新增一個按鈕來啟動編輯郵件內容的視窗
        edit_button = QPushButton('編輯內容', self)
        edit_button.setGeometry(50, 150, 200, 30)
        edit_button.clicked.connect(self.edit_message_button_clicked)

        self.smtp_server = 'smtp.gmail.com'
        self.smtp_server_port = 587
        self.smtp_server_account = 'moccapower.tw@gmail.com'
        self.smtp_server_password = 'rqydtumscyxeajxm'
        self.server = smtplib.SMTP(self.smtp_server, self.smtp_server_port)
        self.server_res = self.server.ehlo()
        self.smtp_ttls = self.server.starttls()
        self.smtp_login = self.server.login(self.smtp_server_account, self.smtp_server_password)

        self.info = None 
        self.keys = None

    @pyqtSlot()
    def edit_message_button_clicked(self):
        if self.mail_type == 0:
            return
        print(self.mail_type)
        self.text_edit.setHtml(self.mail_format[self.mail_type-1])
        edit_dialog = EditMessageDialog(self)
        edit_dialog.set_text(self.text_edit.toHtml())
        result = edit_dialog.exec_()

        if result == QDialog.Accepted:
            edited_text = edit_dialog.get_text()
            self.mail_format[self.mail_type-1]  = edited_text
            self.text_edit.setHtml(self.mail_format[self.mail_type-1])
            
    @pyqtSlot()
    def import_file(self):
        try:
            file_dialog = QFileDialog()
            file_path, _ = file_dialog.getOpenFileName()
            file_name = file_path.split('/')[-1]
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active
            info = {}
            keys = []  # Move keys initialization outside the loop
            for cell in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
                for value in cell:
                    keys.append(value)

            for i in range(2, sheet.max_row + 1):
                row_data = {}
                for j, key in enumerate(keys):
                    row_data[key] = str(sheet.cell(row=i, column=j + 1).value)
                info[i] = row_data

            self.keys = keys
            self.info = info

            self.import_status_label.setText(f'{file_name}匯入成功！')
        except Exception as e:
            logger.exception('匯入文件時發生錯誤：')
            self.import_status_label.setText('匯入 Excel 檔案時發生錯誤')


    @pyqtSlot()
    def send_email_button_clicked(self):
        if not self.info:
            return
        from_mail = self.smtp_server_account
        suc = 0
        fail = 0
        for i in self.info.values():
            if i['Email'] == 'None' or i['信件類型'] == 'None':
                continue
            i['搭車日期'] = i['搭車日期'].split()[0]
            print(i['Email'])

            to_mail = {
                'name': i['Email'],
                'addr': i['Email']
            }
            mail_type = int(i['信件類型'])
            self.text_edit.setHtml(self.mail_format[mail_type-1])
            to_message = self.text_edit.toHtml()
            for k in self.keys:
                if k is not None:
                    form = '{' + k + '}'
                    to_message = to_message.replace(form, i[k])  # Assign the result back to to_message
            text = MIMEText(str(to_message), 'html', 'utf-8')  # Convert to Python str before MIMEText
            text['Subject'] = 'Mocca Service摩克動力機場接送 駕駛資料通知 Driver details 訂單編號: ' + i['訂單編號']
            text['From'] = 'Moccapower'
            text['To'] = to_mail['name']
            text = text.as_string()
            try:
                self.server.sendmail(from_mail, to_mail['addr'], text)
                logger.info('寄信成功 - 訂單編號:' + i['訂單編號']+ f' - 郵件地址：{to_mail["addr"]}' + f' - 訂單類型: {i["信件類型"]}')
                suc += 1
            except Exception as e:
                logger.error('寄信失敗 - 訂單編號:' + i['訂單編號']+ f' - 郵件地址：{to_mail["addr"]} ' + f' - 訂單類型: {i["信件類型"]}')
                fail += 1

        self.import_status_label.setText(f'共寄出{suc+fail}件，成功{suc}，失敗{fail}')

        # self.server.quit()

    @pyqtSlot()
    def send_report_email_button_clicked(self):
        self.mail_format_lable.setText('以選擇 回報駕駛')
        self.mail_type = 1


    @pyqtSlot()
    def send_welcome_ch_email_button_clicked(self):
        self.mail_format_lable.setText('以選擇 歡迎信(中文)')
        self.mail_type = 2

    @pyqtSlot()
    def send_welcome_eng_email_button_clicked(self):
        self.mail_format_lable.setText('以選擇 歡迎信(英文)')
        self.mail_type = 3

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = EmailWindow()
    window.show()
    sys.exit(app.exec_())
